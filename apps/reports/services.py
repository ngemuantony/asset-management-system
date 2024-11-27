import os
from django.template.loader import render_to_string
from django.conf import settings
from weasyprint import HTML
from openpyxl import Workbook
import csv
from datetime import datetime

class ReportGenerator:
    """
    Service class for generating reports in various formats
    """
    
    @staticmethod
    def generate_pdf(template_name, context, output_file):
        """Generate PDF report using WeasyPrint"""
        html_string = render_to_string(template_name, context)
        html = HTML(string=html_string)
        html.write_pdf(output_file)
        return output_file

    @staticmethod
    def generate_excel(data, headers, output_file):
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item.get(key, ''))
        
        wb.save(output_file)
        return output_file

    @staticmethod
    def generate_csv(data, headers, output_file):
        """Generate CSV report"""
        with open(output_file, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        return output_file

class ReportScheduler:
    """
    Service class for scheduling report generation
    """
    
    @staticmethod
    def schedule_report(report_template, parameters, schedule_time):
        """Schedule a report for generation"""
        # Implementation depends on chosen task queue (Celery, Django-Q, etc.)
        pass

    @staticmethod
    def cancel_scheduled_report(report_id):
        """Cancel a scheduled report"""
        pass
